"""
io_excel.py
-----------
Tudo que entra e sai por ARQUIVO (seções 3.6 e 3.7 da especificação):

  - importar_arquivo()  -> lê .csv ou .xlsx, valida linha a linha e SÓ
                           libera para gravar se nenhuma linha tiver erro
  - gerar_template()    -> monta o template_lancamentos.xlsx com a aba
                           "Lançamentos" e a aba "Instruções"
  - exportar_excel()    -> exporta Dados / Indicadores / Previsão em 3
                           abas, com semáforo por cor de célula

A validação do formulário da tela de lançamentos também usa as funções
daqui (validar_linha), para as regras serem exatamente as mesmas nos
dois caminhos de entrada.
"""

import csv
import io
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from dados import CAMPOS_LANCAMENTO, NOMES_CAMPOS

# Cores do semáforo no Excel (formato ARGB do openpyxl)
CORES_EXCEL = {
    "saudavel": "C6EFCE",  # verde claro
    "atencao": "FFEB9C",   # amarelo claro
    "alerta": "FFD8B0",    # laranja claro
    "critico": "FFC7CE",   # vermelho claro
}

# O mês precisa estar no formato AAAA-MM (ex.: 2026-07)
FORMATO_MES = re.compile(r"^\d{4}-(0[1-9]|1[0-2])$")


# ---------------------------------------------------------------------------
# Validação (usada pela importação E pelo formulário)
# ---------------------------------------------------------------------------

def interpretar_numero(texto):
    """
    Converte texto em número aceitando o jeito brasileiro de escrever:
    "1.234,56" ou "1234.56" ou "1234". Devolve float ou levanta ValueError.
    """
    texto = str(texto).strip().replace("R$", "").replace(" ", "")
    if "," in texto:
        # Formato brasileiro: ponto é milhar, vírgula é decimal
        texto = texto.replace(".", "").replace(",", ".")
    return float(texto)


def interpretar_sim_nao(texto):
    """Converte SIM/NÃO (e variações: S, N, 1, 0) em 1 ou 0."""
    texto = str(texto).strip().upper()
    if texto in ("SIM", "S", "1", "1.0", "TRUE", "VERDADEIRO"):
        return 1
    if texto in ("NAO", "NÃO", "N", "0", "0.0", "FALSE", "FALSO"):
        return 0
    raise ValueError(f"esperado SIM ou NÃO, veio '{texto}'")


def validar_linha(linha_crua):
    """
    Valida UMA linha de lançamento (dicionário de textos, como vem do CSV
    ou do formulário) e devolve (lancamento_convertido, lista_de_erros).

    Regras aplicadas (seção 3.6):
      - campos obrigatórios não podem estar vazios
      - mês no formato AAAA-MM
      - números precisam ser números; inteiros precisam ser inteiros
      - valores em R$ não podem ser negativos (exceto o saldo de caixa,
        que pode ficar negativo se a conta estiver no vermelho)
    """
    lancamento = {}
    erros = []

    for campo in CAMPOS_LANCAMENTO:
        nome, tipo = campo["nome"], campo["tipo"]
        bruto = linha_crua.get(nome)
        vazio = bruto is None or str(bruto).strip() == ""

        if vazio:
            if campo["obrigatorio"]:
                erros.append(f"campo obrigatório '{nome}' está vazio")
            else:
                lancamento[nome] = None  # opcional em branco: tudo bem
            continue

        try:
            if tipo == "texto":  # só o campo "mes" é texto
                valor = str(bruto).strip()
                if not FORMATO_MES.match(valor):
                    erros.append(f"'{nome}' deve estar no formato AAAA-MM (veio '{valor}')")
                    continue
            elif tipo == "sim_nao":
                valor = interpretar_sim_nao(bruto)
            elif tipo == "inteiro":
                valor = int(interpretar_numero(bruto))
            else:  # "valor" em R$
                valor = interpretar_numero(bruto)
                if valor < 0 and nome != "saldo_caixa_final":
                    erros.append(f"'{nome}' não pode ser negativo (veio {valor})")
                    continue
            lancamento[nome] = valor
        except ValueError:
            erros.append(f"'{nome}' tem valor inválido ('{bruto}')")

    return lancamento, erros


def validar_linhas(linhas_cruas):
    """
    Valida a planilha inteira e devolve (lancamentos_validos, relatorio_de_erros).

    O relatório é uma lista de {"linha": nº, "erros": [...]} — uma entrada
    por linha problemática. Se o relatório NÃO estiver vazio, nada deve
    ser gravado (regra da especificação: relatório de erros antes de
    gravar qualquer coisa).
    """
    lancamentos = []
    relatorio = []
    meses_vistos = set()

    for numero, linha_crua in enumerate(linhas_cruas, start=2):  # linha 1 = cabeçalho
        lancamento, erros = validar_linha(linha_crua)

        # Mês duplicado dentro do próprio arquivo
        mes = lancamento.get("mes")
        if mes and mes in meses_vistos:
            erros.append(f"mês '{mes}' aparece mais de uma vez no arquivo")
        if mes:
            meses_vistos.add(mes)

        if erros:
            relatorio.append({"linha": numero, "erros": erros})
        else:
            lancamentos.append(lancamento)

    return lancamentos, relatorio


# ---------------------------------------------------------------------------
# Importação de arquivo (.csv ou .xlsx)
# ---------------------------------------------------------------------------

def _ler_csv(conteudo_bytes):
    """Lê os bytes de um CSV e devolve lista de dicionários de texto."""
    # Tentamos UTF-8 primeiro; se falhar, Latin-1 (comum em Excel brasileiro)
    try:
        texto = conteudo_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        texto = conteudo_bytes.decode("latin-1")
    # Detecta o separador: Excel em português costuma salvar CSV com ";"
    separador = ";" if texto.splitlines()[0].count(";") > texto.splitlines()[0].count(",") else ","
    return list(csv.DictReader(io.StringIO(texto), delimiter=separador))


def _ler_xlsx(conteudo_bytes):
    """Lê os bytes de um .xlsx (primeira aba) e devolve lista de dicionários."""
    planilha = load_workbook(io.BytesIO(conteudo_bytes), data_only=True).worksheets[0]
    linhas = list(planilha.iter_rows(values_only=True))
    if not linhas:
        return []
    cabecalho = [str(celula).strip() if celula is not None else "" for celula in linhas[0]]
    resultado = []
    for linha in linhas[1:]:
        if all(celula is None for celula in linha):
            continue  # pula linhas totalmente vazias no fim da planilha
        resultado.append(dict(zip(cabecalho, linha)))
    return resultado


def importar_arquivo(conteudo_bytes, nome_arquivo):
    """
    Porta de entrada da importação: recebe os bytes e o nome do arquivo
    enviado, escolhe o leitor certo (.csv ou .xlsx), confere se as colunas
    obrigatórias existem e valida linha a linha.

    Devolve (lancamentos_validos, relatorio_de_erros). Quem chama (app.py)
    só grava se o relatório vier vazio.
    """
    if nome_arquivo.lower().endswith(".xlsx"):
        linhas_cruas = _ler_xlsx(conteudo_bytes)
    elif nome_arquivo.lower().endswith(".csv"):
        linhas_cruas = _ler_csv(conteudo_bytes)
    else:
        return [], [{"linha": "-", "erros": ["formato não suportado: envie um .csv ou .xlsx"]}]

    if not linhas_cruas:
        return [], [{"linha": "-", "erros": ["o arquivo está vazio"]}]

    # Confere o cabeçalho: todas as colunas obrigatórias precisam existir
    colunas_do_arquivo = set(linhas_cruas[0].keys())
    obrigatorias = [c["nome"] for c in CAMPOS_LANCAMENTO if c["obrigatorio"]]
    faltando = [nome for nome in obrigatorias if nome not in colunas_do_arquivo]
    if faltando:
        return [], [{"linha": 1, "erros": [f"colunas obrigatórias ausentes no cabeçalho: {', '.join(faltando)}"]}]

    return validar_linhas(linhas_cruas)


# ---------------------------------------------------------------------------
# Template oficial de lançamentos
# ---------------------------------------------------------------------------

def gerar_template():
    """
    Gera o template_lancamentos.xlsx em memória (BytesIO), com:
      - aba "Lançamentos": cabeçalho com as 13 colunas + 1 linha de exemplo
      - aba "Instruções": explicação de cada campo
    """
    pasta_de_trabalho = Workbook()

    aba_dados = pasta_de_trabalho.active
    aba_dados.title = "Lançamentos"
    aba_dados.append(NOMES_CAMPOS)
    for celula in aba_dados[1]:
        celula.font = Font(bold=True)
    # Linha de exemplo (o usuário substitui pelos dados reais)
    aba_dados.append(["2026-06", 6500, 1800, 2200, 2000, 3100, 450, 5400,
                      120, 85, 300, "SIM", 900])

    aba_instrucoes = pasta_de_trabalho.create_sheet("Instruções")
    aba_instrucoes.append(["Campo", "O que preencher", "Obrigatório?"])
    for celula in aba_instrucoes[1]:
        celula.font = Font(bold=True)
    for campo in CAMPOS_LANCAMENTO:
        aba_instrucoes.append([
            campo["nome"],
            campo["rotulo"],
            "SIM" if campo["obrigatorio"] else "não",
        ])
    aba_instrucoes.append([])
    aba_instrucoes.append(["Dica: um lançamento por mês, mês no formato AAAA-MM (ex.: 2026-06)."])
    aba_instrucoes.append(["Valores em reais podem usar vírgula (ex.: 1234,56)."])
    aba_instrucoes.column_dimensions["A"].width = 26
    aba_instrucoes.column_dimensions["B"].width = 42

    memoria = io.BytesIO()
    pasta_de_trabalho.save(memoria)
    memoria.seek(0)
    return memoria


# ---------------------------------------------------------------------------
# Exportação com 3 abas (Dados / Indicadores / Previsão)
# ---------------------------------------------------------------------------

# Colunas da aba Indicadores: (título, chave no dicionário de indicadores)
COLUNAS_INDICADORES = [
    ("Margem líquida", "margem_liquida"),
    ("Crescimento receita", "crescimento_receita"),
    ("Fôlego de caixa (meses)", "runway_meses"),
    ("Comprometimento dívidas", "comprometimento_dividas"),
    ("Uso do limite MEI", "uso_limite_mei"),
    ("Inadimplência", "inadimplencia"),
    ("Concentração clientes", "concentracao_clientes"),
    ("Compras/receita", "compras_receita"),
    ("Ticket médio (R$)", "ticket_medio"),
    ("Regularidade DAS", "regularidade_das"),
]


def exportar_excel(lancamentos, indicadores_por_mes, scores_por_mes, previsao):
    """
    Monta o Excel de exportação em memória (BytesIO) com 3 abas:
      - Dados:       os lançamentos como foram digitados
      - Indicadores: os 10 KPIs de cada mês + score + classe (célula do
                     score pintada com a cor do semáforo)
      - Previsão:    cenários dos próximos meses (ou o aviso de dados
                     insuficientes)
    """
    pasta_de_trabalho = Workbook()

    # --- Aba 1: Dados -------------------------------------------------------
    aba_dados = pasta_de_trabalho.active
    aba_dados.title = "Dados"
    aba_dados.append(NOMES_CAMPOS)
    for celula in aba_dados[1]:
        celula.font = Font(bold=True)
    for lancamento in lancamentos:
        linha = []
        for nome in NOMES_CAMPOS:
            valor = lancamento.get(nome)
            if nome == "das_pago_em_dia":
                valor = "SIM" if valor == 1 else "NÃO"
            linha.append(valor)
        aba_dados.append(linha)

    # --- Aba 2: Indicadores + score com semáforo ----------------------------
    aba_indicadores = pasta_de_trabalho.create_sheet("Indicadores")
    aba_indicadores.append(
        ["Mês"] + [titulo for titulo, _ in COLUNAS_INDICADORES] + ["Score", "Classe"]
    )
    for celula in aba_indicadores[1]:
        celula.font = Font(bold=True)

    for indicadores, resultado_score in zip(indicadores_por_mes, scores_por_mes):
        linha = [indicadores["mes"]]
        for _, chave in COLUNAS_INDICADORES:
            valor = indicadores.get(chave)
            linha.append(round(valor, 4) if valor is not None else None)
        linha += [resultado_score["score"], resultado_score["classe"]]
        aba_indicadores.append(linha)

        # Pinta as células de Score e Classe com a cor da classe (semáforo)
        cor = PatternFill("solid", fgColor=CORES_EXCEL[resultado_score["nivel"]])
        numero_linha = aba_indicadores.max_row
        aba_indicadores.cell(numero_linha, len(COLUNAS_INDICADORES) + 2).fill = cor
        aba_indicadores.cell(numero_linha, len(COLUNAS_INDICADORES) + 3).fill = cor

    # --- Aba 3: Previsão -----------------------------------------------------
    aba_previsao = pasta_de_trabalho.create_sheet("Previsão")
    if previsao.get("disponivel"):
        aba_previsao.append(["Mês", "Receita (pessimista)", "Receita (base)", "Receita (otimista)",
                             "Caixa (base)", "Score (pessimista)", "Score (base)", "Score (otimista)"])
        for celula in aba_previsao[1]:
            celula.font = Font(bold=True)
        for posicao, mes in enumerate(previsao["meses"]):
            aba_previsao.append([
                mes,
                previsao["receita"]["pessimista"][posicao],
                previsao["receita"]["base"][posicao],
                previsao["receita"]["otimista"][posicao],
                previsao["caixa"]["base"][posicao],
                previsao["score"]["pessimista"][posicao],
                previsao["score"]["base"][posicao],
                previsao["score"]["otimista"][posicao],
            ])
        aba_previsao.append([])
        aba_previsao.append([f"Modelo: {previsao['modelo']}"])
        aba_previsao.append([f"Erro padrão da receita (RMSE): {previsao['erro_receita']:.2f}"])
    else:
        aba_previsao.append([previsao.get("motivo", "Previsão indisponível.")])

    memoria = io.BytesIO()
    pasta_de_trabalho.save(memoria)
    memoria.seek(0)
    return memoria
